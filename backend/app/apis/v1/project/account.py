from uuid import UUID
from typing import List

from fastapi import APIRouter, Query, Body, HTTPException, Path, Depends

from app.schemas.project.account import Create, Update, Out, OutList, StatsOut, BalanceStats, VariableStats, StatsData
from app.crud.project.account import project_account_crud
from app.utils.time_tool import parse_time
from app.schemas.base import BaseOut


from app.apis.deps import get_current_user, get_admin_user, get_gm_user

app = APIRouter()


@app.post("", response_model=Out, description="创建项目账号", summary="创建项目账号")
async def post(
    item: Create = Body(..., description="创建数据"),
    current_user: dict = Depends(get_current_user)
):
    """
    创建项目账号记录
    """
    try:
        return await project_account_crud.create(item)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/stats", response_model=StatsOut, description="统计项目账号数据", summary="统计项目账号数据")
async def get_stats(
    project_id: UUID = Query(..., description="项目ID（必填）"),
    account: str | None = Query(None, description="账号"),
    status: int | None = Query(None, description="账号状态"),
    account_type: int | None = Query(None, description="账号类型"),
    create_time_start: str | int | None = Query(
        None,
        description="创建时间开始 (支持 YYYY-MM-DD / YYYY-MM-DD HH:mm:ss / 13位时间戳)",
    ),
    create_time_end: str | int | None = Query(
        None,
        description="创建时间结束 (支持 YYYY-MM-DD / YYYY-MM-DD HH:mm:ss / 13位时间戳)",
    ),
    update_time_start: str | int | None = Query(
        None,
        description="更新时间开始 (支持 YYYY-MM-DD / YYYY-MM-DD HH:mm:ss / 13位时间戳)",
    ),
    update_time_end: str | int | None = Query(
        None,
        description="更新时间结束 (支持 YYYY-MM-DD / YYYY-MM-DD HH:mm:ss / 13位时间戳)",
    ),
    current_user: dict = Depends(get_current_user)
):
    """
    统计项目账号的余额和变动数据
    返回：最高分、最低分、平均分、总分（余额和变动）
    
    权限控制：
    - ADMIN/GM: 可以统计所有项目
    - IT/MANUAL: 只能统计分配给自己的项目
    """
    try:
        from app.utils.data_permission import filter_by_user_projects
        from decimal import Decimal
        
        # 获取用户ID
        user_id = current_user.get('user_id') or current_user.get('id')
        
        # 根据用户权限过滤项目
        user_project_ids = await filter_by_user_projects(user_id)
        
        # 检查用户是否有权限访问该项目
        if user_project_ids is not None:
            if str(project_id) not in user_project_ids:
                raise HTTPException(status_code=403, detail="没有权限访问该项目")
        
        # 调用CRUD层的统计方法
        stats = await project_account_crud.get_stats(
            project_id=project_id,
            account=account,
            status=status,
            account_type=account_type,
            create_time_start=parse_time(create_time_start),
            create_time_end=parse_time(create_time_end, True),
            update_time_start=parse_time(update_time_start),
            update_time_end=parse_time(update_time_end, True),
        )
        
        # 转换Decimal为float，保留6位小数
        def decimal_to_float(value):
            if value is None:
                return 0.0
            return float(Decimal(str(value)))
        
        # 构建响应数据
        return StatsOut(
            code=1,
            message="成功",
            data=StatsData(
                total_count=stats.get("total_count", 0),
                balance=BalanceStats(
                    max=decimal_to_float(stats.get("max_balance")),
                    min=decimal_to_float(stats.get("min_balance")),
                    avg=decimal_to_float(stats.get("avg_balance")),
                    sum=decimal_to_float(stats.get("sum_balance")),
                ),
                variable=VariableStats(
                    max=decimal_to_float(stats.get("max_variable")),
                    min=decimal_to_float(stats.get("min_variable")),
                    avg=decimal_to_float(stats.get("avg_variable")),
                    sum=decimal_to_float(stats.get("sum_variable")),
                )
            )
        )
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/export-all-stats", description="导出所有项目统计数据为Excel", summary="导出所有项目统计数据")
async def export_all_stats(
    current_user: dict = Depends(get_gm_user)
):
    """
    导出所有项目的统计数据为Excel文件
    
    权限控制：
    - 仅 ADMIN/GM 可以导出
    
    返回：
    - Excel文件流
    """
    try:
        from fastapi.responses import StreamingResponse
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from app.models.project import ProjectInfo, ProjectStatus
        from decimal import Decimal
        import datetime
        from urllib.parse import quote
        
        # 获取所有项目（预加载用户关联）
        projects = await ProjectInfo.all().prefetch_related('users')
        
        if not projects:
            raise HTTPException(status_code=404, detail="没有项目数据")
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "项目统计汇总"
        
        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 项目状态映射
        status_map = {
            ProjectStatus.NORMAL: "正常",
            ProjectStatus.NOT_WRITTEN: "未编写",
            ProjectStatus.WRITING: "编写中",
            ProjectStatus.ENDED: "项目结束",
            ProjectStatus.RUNAWAY: "项目跑路",
            ProjectStatus.MAINTENANCE: "项目维护",
            ProjectStatus.UNASSIGNED: "未分配",
            ProjectStatus.ACCOUNT_NOT_SUPPORT: "账号不支持",
            ProjectStatus.IP_NOT_SUPPORT: "IP不支持",
        }
        
        # 定义表头（添加项目状态列）
        headers = [
            "项目名称", "项目状态", "项目ID", "所属用户", "账号数量",
            "余额最高分", "余额最低分", "余额平均分", "余额总分",
            "变动最高分", "变动最低分", "变动平均分", "变动总分"
        ]
        
        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 设置列宽（调整项目状态列宽度）
        column_widths = [20, 12, 38, 25, 12, 12, 12, 12, 12, 12, 12, 12, 12]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_num)].width = width
        
        # 转换Decimal为float
        def decimal_to_float(value):
            if value is None:
                return 0.0
            return float(Decimal(str(value)))
        
        # 遍历所有项目，获取统计数据
        row_num = 2
        for project in projects:
            try:
                # 获取该项目的统计数据
                stats = await project_account_crud.get_stats(
                    project_id=project.id,
                    account=None,
                    status=None,
                    account_type=None,
                    create_time_start=None,
                    create_time_end=None,
                    update_time_start=None,
                    update_time_end=None,
                )
                
                # 获取项目关联的用户昵称（多个用户用逗号分隔）
                user_nicknames = []
                if project.users:
                    for user in project.users:
                        if user.nickname:
                            user_nicknames.append(user.nickname)
                        else:
                            user_nicknames.append(user.email)  # 如果没有昵称，使用邮箱
                
                users_str = ", ".join(user_nicknames) if user_nicknames else "未分配"
                
                # 获取项目状态文本
                status_text = status_map.get(project.status, "未知")
                
                # 写入数据行
                row_data = [
                    project.name,
                    status_text,  # 项目状态
                    str(project.id),
                    users_str,  # 所属用户
                    stats.get("total_count", 0),
                    round(decimal_to_float(stats.get("max_balance")), 2),
                    round(decimal_to_float(stats.get("min_balance")), 2),
                    round(decimal_to_float(stats.get("avg_balance")), 2),
                    round(decimal_to_float(stats.get("sum_balance")), 2),
                    round(decimal_to_float(stats.get("max_variable")), 2),
                    round(decimal_to_float(stats.get("min_variable")), 2),
                    round(decimal_to_float(stats.get("avg_variable")), 2),
                    round(decimal_to_float(stats.get("sum_variable")), 2),
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    # 前4列左对齐，其他列居中
                    cell.alignment = Alignment(horizontal="left" if col_num <= 4 else "center", vertical="center")
                
                row_num += 1
            except Exception as e:
                # 如果某个项目统计失败，记录错误但继续处理其他项目
                print(f"项目 {project.name} 统计失败: {str(e)}")
                continue
        
        # 保存到内存
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # 生成文件名（包含当前日期时间）
        now = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"project_stats_{now}.xlsx"  # 使用英文文件名避免编码问题
        
        # 返回文件流
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@app.get("/{id}", response_model=Out, description="获取项目账号", summary="获取项目账号")
async def get(
    id: UUID = Path(..., description="ID"),
    current_user: dict = Depends(get_current_user)
):
    """
    获取单个项目账号记录
    - 敏感字段（private_key、mnemonic）根据权限决定是否解密
    - 项目所属人和ADMIN可以看到解密后的数据
    - 其他用户看到加密数据
    """
    try:
        # 获取用户ID和角色
        user_id = current_user.get('user_id') or current_user.get('id')
        user_roles = current_user.get('roles', [])
        
        obj = await project_account_crud.get(id, user_id=str(user_id), user_roles=user_roles)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    if not obj:
        raise HTTPException(status_code=404, detail="数据不存在")
    return obj


@app.get("", response_model=OutList, description="获取项目账号列表", summary="获取项目账号列表")
async def gets(
    account: str | None = Query(None, description="账号"),
    status: int | None = Query(None, description="账号状态"),
    account_type: int | None = Query(None, description="账号类型"),
    project_id: UUID | None = Query(None, description="所属项目ID"),
    server_id: UUID | None = Query(None, description="关联服务器信息ID"),
    order_by: str | None = Query(
        "-create_time",
        description="排序字段",
        pattern="^(?:-)?(?:id|account|status|create_time|update_time)$",
    ),
    res_count: bool = Query(False, description="是否返回总数"),
    create_time_start: str | int | None = Query(
        None,
        description="创建时间开始 (支持 YYYY-MM-DD / YYYY-MM-DD HH:mm:ss / 13位时间戳)",
    ),
    create_time_end: str | int | None = Query(
        None,
        description="创建时间结束 (支持 YYYY-MM-DD / YYYY-MM-DD HH:mm:ss / 13位时间戳)",
    ),
    update_time_start: str | int | None = Query(
        None,
        description="更新时间开始 (支持 YYYY-MM-DD / YYYY-MM-DD HH:mm:ss / 13位时间戳)",
    ),
    update_time_end: str | int | None = Query(
        None,
        description="更新时间结束 (支持 YYYY-MM-DD / YYYY-MM-DD HH:mm:ss / 13位时间戳)",
    ),
    page: int = Query(1, ge=1, description="页码"),
    limit: int = Query(10, ge=1, le=1000, description="每页数量"),
    current_user: dict = Depends(get_current_user)
):
    """
    分页查询项目账号列表
    根据用户角色返回不同的数据：
    - ADMIN/GM: 返回所有项目的账号，敏感字段解密
    - IT/MANUAL: 只返回分配给该用户的项目的账号，敏感字段解密
    - 其他用户：敏感字段保持加密状态
    """
    try:
        from app.utils.data_permission import filter_by_user_projects
        
        # 获取用户ID和角色
        user_id = current_user.get('user_id') or current_user.get('id')
        user_roles = current_user.get('roles', [])
        
        # 根据用户权限过滤项目
        user_project_ids = await filter_by_user_projects(user_id)
        
        # 如果指定了project_id，需要检查用户是否有权限访问该项目
        if project_id and user_project_ids is not None:
            if str(project_id) not in user_project_ids:
                # 用户没有权限访问该项目
                from app.schemas.project.account import OutList
                return OutList(message='成功', count=0, num=0, items=[])
        
        return await project_account_crud.get_multi(
            account=account,
            status=status,
            account_type=account_type,
            project_id=project_id,
            server_id=server_id,
            order_by=order_by or "-create_time",
            create_time_start=parse_time(create_time_start),
            create_time_end=parse_time(create_time_end, True),
            update_time_start=parse_time(update_time_start),
            update_time_end=parse_time(update_time_end, True),
            page=page,
            limit=limit,
            res_count=res_count,
            user_project_ids=user_project_ids,
            user_id=str(user_id),
            user_roles=user_roles,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/{id}", response_model=Out, description="更新项目账号", summary="更新项目账号")
async def put(
    id: UUID = Path(..., description="主键ID"),
    item: Update = Body(..., description="更新数据"),
    current_user: dict = Depends(get_current_user)
):
    """
    部分更新项目账号，只更新传入的非空字段
    """
    try:
        return await project_account_crud.update(id, item)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/{id}", response_model=BaseOut, description="删除项目账号", summary="删除项目账号")
async def delete(
    id: UUID = Path(..., description="主键ID"),
    current_user: dict = Depends(get_current_user)
):
    """
    删除项目账号
    - 用户可以删除自己有权限的项目下的账号
    - 管理员可以删除所有账号
    """
    try:
        from app.utils.data_permission import filter_by_user_projects, has_resource_access
        
        user_id = current_user['user_id']
        user_roles = current_user.get('roles', [])
        
        # 检查是否有访问项目的权限（项目账号属于项目资源）
        if not has_resource_access(user_roles, 'project'):
            raise HTTPException(status_code=403, detail="没有访问项目的权限")
        
        # 获取要删除的账号
        account = await project_account_crud.get(id)
        if not account:
            raise HTTPException(status_code=404, detail="账号不存在")
        
        # 检查项目权限
        allowed_project_ids = await filter_by_user_projects(user_id)
        
        # 如果不是全局权限（ADMIN/GM），检查是否有该项目的权限
        if allowed_project_ids is not None:
            if str(account.project_id) not in [str(pid) for pid in allowed_project_ids]:
                raise HTTPException(status_code=403, detail="没有权限删除该项目下的账号")
        
        # 执行删除
        await project_account_crud.delete(id)
        return BaseOut(message="成功", count=1)
    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/upsert", response_model=BaseOut, description="创建或更新项目账号（使用Redis队列）", summary="创建或更新项目账号")
async def post_or_put(
    item: Create = Body(..., description="创建或更新数据"),
    current_user: dict = Depends(get_current_user)
):
    """
    创建或更新项目账号（使用Redis队列异步处理）
    根据 account 和 project_id 判断是否存在：
    - 如果存在，只更新传入的非空字段
    - 如果不存在，创建新记录
    
    数据会被添加到Redis队列，由后台worker异步处理
    """
    try:
        from app.utils.project_account_queue import project_account_queue
        from app.core.settings import REDIS_ENABLED
        
        if not REDIS_ENABLED:
            raise HTTPException(status_code=503, detail="Redis未启用，无法使用队列处理功能")
        
        # 转换为字典，使用mode='json'确保UUID和Enum都能被序列化
        data = item.model_dump(mode='json')
        
        # 添加到队列
        if await project_account_queue.add_to_queue(data):
            # 获取当前队列大小
            queue_size = await project_account_queue.get_queue_size()
            return BaseOut(
                message=f"成功添加到队列，当前队列大小: {queue_size}",
                count=1
            )
        else:
            raise HTTPException(status_code=500, detail="添加到队列失败")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/batch-upsert", response_model=BaseOut, description="批量创建或更新项目账号（使用Redis队列）", summary="批量创建或更新项目账号")
async def batch_upsert(
    items: List[Create] = Body(..., description="批量创建或更新数据"),
    current_user: dict = Depends(get_current_user)
):
    """
    批量创建或更新项目账号
    使用Redis队列异步处理，避免数据库压力和接口长时间占用
    """
    try:
        from app.utils.project_account_queue import project_account_queue
        from app.core.settings import REDIS_ENABLED
        
        if not REDIS_ENABLED:
            raise HTTPException(status_code=503, detail="Redis未启用，无法使用批量处理功能")
        
        # 将数据添加到Redis队列
        success_count = 0
        fail_count = 0
        
        for item in items:
            # 转换为字典，使用mode='json'确保UUID和Enum都能被序列化
            data = item.model_dump(mode='json')
            # 添加到队列
            if await project_account_queue.add_to_queue(data):
                success_count += 1
            else:
                fail_count += 1
        
        # 获取当前队列大小
        queue_size = await project_account_queue.get_queue_size()
        
        return BaseOut(
            message=f"成功添加 {success_count} 条数据到队列，失败 {fail_count} 条，当前队列大小: {queue_size}",
            count=success_count
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))




